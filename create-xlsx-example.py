#!/usr/bin/env python3
"""
Create professional XLSX file for PhotoCertif batch certification
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook with multiple sheets
wb = Workbook()

# ============================================
# Sheet 1: Data with Examples
# ============================================
ws_data = wb.active
ws_data.title = "Certification Data"

# Headers
headers = [
    "fileUrl",
    "title", 
    "description",
    "cert_name",
    "cert_symbol",
    "cert_description",
    "cert_owner",
    "external_url",
    "twitter_url",
    "discord_url",
    "instagram_url",
    "telegram_url",
    "medium_url",
    "wiki_url",
    "youtube_url"
]

# Style for headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Example data rows
examples = [
    {
        "fileUrl": "https://drive.google.com/uc?id=1abc123def456&export=download",
        "title": "My Artwork 2025",
        "description": "Digital art certified on blockchain",
        "cert_name": "MyArt2025",
        "cert_symbol": "MYART",
        "cert_description": "Original digital artwork certified with AI authenticity analysis",
        "cert_owner": "John Doe Artist",
        "external_url": "https://myartportfolio.com",
        "twitter_url": "https://x.com/johnartist",
        "discord_url": "https://discord.gg/myartcommunity",
        "instagram_url": "https://instagram.com/johnartist",
        "telegram_url": "https://t.me/johnartistchannel",
        "medium_url": "https://medium.com/@johnartist",
        "wiki_url": "https://docs.myartproject.com",
        "youtube_url": "https://youtube.com/@johnartist"
    },
    {
        "fileUrl": "https://www.dropbox.com/s/xyz789/contract.pdf?dl=1",
        "title": "Legal Contract Q1 2025",
        "description": "Important legal document certified",
        "cert_name": "LegalDoc2025",
        "cert_symbol": "LEGAL",
        "cert_description": "Official contract document with timestamp certification",
        "cert_owner": "Acme Corporation",
        "external_url": "https://acmecorp.com",
        "twitter_url": "",
        "discord_url": "",
        "instagram_url": "",
        "telegram_url": "",
        "medium_url": "",
        "wiki_url": "",
        "youtube_url": ""
    },
    {
        "fileUrl": "https://cdn.example.com/photos/sunset.jpg",
        "title": "Sunset Photography",
        "description": "Professional photography work",
        "cert_name": "Sunset2025",
        "cert_symbol": "PHOTO",
        "cert_description": "Original photograph captured on location",
        "cert_owner": "Jane Smith Photo",
        "external_url": "https://janesmithphotography.com",
        "twitter_url": "https://x.com/janesmith",
        "discord_url": "",
        "instagram_url": "",
        "telegram_url": "https://t.me/janesmithphoto",
        "medium_url": "",
        "wiki_url": "https://wiki.janesmith.com",
        "youtube_url": ""
    }
]

# Write example rows
row_fill_light = PatternFill(start_color="E7E6F4", end_color="E7E6F4", fill_type="solid")
row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for row_num, example in enumerate(examples, 2):
    fill = row_fill_light if row_num % 2 == 0 else row_fill_white
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=row_num, column=col_num, value=example.get(header, ""))
        cell.border = thin_border
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=False)

# Add empty rows for user to fill
for row_num in range(5, 11):  # Add 6 empty rows
    for col_num in range(1, len(headers) + 1):
        cell = ws_data.cell(row=row_num, column=col_num, value="")
        cell.border = thin_border

# Adjust column widths
column_widths = {
    'A': 70,  # fileUrl
    'B': 30,  # title
    'C': 40,  # description
    'D': 20,  # cert_name
    'E': 15,  # cert_symbol
    'F': 50,  # cert_description
    'G': 25,  # cert_owner
    'H': 35,  # external_url
    'I': 35,  # twitter_url
    'J': 40,  # discord_url
    'K': 35,  # instagram_url
    'L': 35,  # telegram_url
    'M': 35,  # medium_url
    'N': 35,  # wiki_url
    'O': 35,  # youtube_url
}

for col, width in column_widths.items():
    ws_data.column_dimensions[col].width = width

# Freeze first row
ws_data.freeze_panes = 'A2'

# ============================================
# Sheet 2: Instructions
# ============================================
ws_instructions = wb.create_sheet("Instructions")

instructions = [
    ["PhotoCertif Batch Certification Guide", ""],
    ["", ""],
    ["Required Fields (7)", ""],
    ["fileUrl", "Public URL to your file (Google Drive, Dropbox, CDN)"],
    ["title", "Title of the certification"],
    ["description", "Description of the content"],
    ["cert_name", "NFT certificate name (no spaces, alphanumeric)"],
    ["cert_symbol", "Short symbol (ex: MYART, max 10 chars)"],
    ["cert_description", "Full description for NFT metadata"],
    ["cert_owner", "Owner/Creator name"],
    ["", ""],
    ["Optional Fields (8) - Social Links", ""],
    ["external_url", "Your main website"],
    ["twitter_url", "Twitter/X profile (https://x.com/username)"],
    ["discord_url", "Discord server invite"],
    ["instagram_url", "Instagram profile"],
    ["telegram_url", "Telegram channel"],
    ["medium_url", "Medium blog"],
    ["wiki_url", "Documentation/Wiki"],
    ["youtube_url", "YouTube channel"],
    ["", ""],
    ["Google Drive URLs", ""],
    ["Format", "https://drive.google.com/uc?id=FILE_ID&export=download"],
    ["How to get FILE_ID", "1. Right-click file → Share → Anyone with link"],
    ["", "2. Copy link: https://drive.google.com/file/d/FILE_ID/view"],
    ["", "3. Extract FILE_ID and use format above"],
    ["", ""],
    ["Dropbox URLs", ""],
    ["Format", "https://www.dropbox.com/s/abc123/file.jpg?dl=1"],
    ["Important", "Change ?dl=0 to ?dl=1 at the end"],
    ["", ""],
    ["Tips", ""],
    ["", "✅ Leave social links empty if not used (don't write N/A)"],
    ["", "✅ Test with 1 row before batch processing"],
    ["", "✅ cert_symbol: max 10 characters, uppercase recommended"],
    ["", "✅ All URLs must be publicly accessible"],
    ["", ""],
    ["n8n Integration", ""],
    ["Step 1", "Upload this file to Google Sheets"],
    ["Step 2", "In n8n: Add Google Sheets node → Read"],
    ["Step 3", "Connect to PhotoCertif workflow"],
    ["Step 4", "Run workflow - each row = 1 certification"],
]

# Write instructions
title_font = Font(size=16, bold=True, color="4472C4")
section_font = Font(size=12, bold=True, color="2F5496")
normal_font = Font(size=11)

for row_num, (col1, col2) in enumerate(instructions, 1):
    cell_a = ws_instructions.cell(row=row_num, column=1, value=col1)
    cell_b = ws_instructions.cell(row=row_num, column=2, value=col2)
    
    # Title formatting
    if row_num == 1:
        cell_a.font = title_font
    # Section headers
    elif col1 and not col2 and col1 not in ["", "Tips", "n8n Integration"]:
        cell_a.font = section_font
    # Subsections
    elif col1 in ["Required Fields (7)", "Optional Fields (8) - Social Links", "Google Drive URLs", "Dropbox URLs", "Tips", "n8n Integration"]:
        cell_a.font = section_font
    else:
        cell_a.font = normal_font
        cell_b.font = normal_font

ws_instructions.column_dimensions['A'].width = 30
ws_instructions.column_dimensions['B'].width = 70

# ============================================
# Save file
# ============================================
filename = "photocertif-batch-example.xlsx"
wb.save(filename)

print(f"✅ Created: {filename}")
print(f"   - Sheet 1: Certification Data (3 examples + 6 empty rows)")
print(f"   - Sheet 2: Instructions & Tips")
print(f"   - {len(headers)} columns with professional formatting")
print(f"   - Ready to import into Google Sheets or use in n8n")
